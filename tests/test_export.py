from openpyxl import load_workbook

from extract_tickets.export import build_csv_bytes, build_excel_bytes

RECORDS = [
    {
        "文件名": "001.pdf",
        "出发站": "威海南海",
        "到达站": "青岛北",
        "车次": "G6989",
        "乘车日期": "2025年10月18日",
        "票价(元)": "110.00",
        "报销状态": "已报销",
    }
]


def test_build_csv_bytes_contains_bom_and_headers():
    csv_bytes = build_csv_bytes(RECORDS).getvalue()

    assert csv_bytes.startswith(b"\xef\xbb\xbf")
    assert bytes("文件名,发票号码", "utf-8") in csv_bytes
    assert b"001.pdf" in csv_bytes


def test_build_excel_bytes_contains_expected_sheets():
    excel_bytes = build_excel_bytes(RECORDS)
    workbook = load_workbook(excel_bytes)

    assert workbook.sheetnames == ["行程明细", "时间统计", "路线分析", "报销汇总"]
    assert workbook["行程明细"]["A2"].value == "001.pdf"
    assert workbook["报销汇总"]["A3"].value == "已报销"
