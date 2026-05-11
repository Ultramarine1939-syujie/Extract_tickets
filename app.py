"""
国铁电子客票 Web 解析工具 - Flask 后端
运行: python app.py
访问: http://localhost:5678

核心解析逻辑已抽取到根目录 parser.py 共用。
"""

import io
import csv
import os
import re
from collections import defaultdict
from flask import Flask, request, jsonify, send_file, render_template

from parser import (
    FIELDNAMES,
    process_pdf_bytes,
    extract_text_from_bytes,
    parse_guotie,
)

FIELDNAMES_EXT = FIELDNAMES + ["报销状态"]


app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100MB 上限


# ── 路由 ──────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/parse", methods=["POST"])
def api_parse():
    """批量解析接口（一次性上传所有文件）。"""
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "未收到文件"}), 400

    records = []
    for f in files:
        if not f.filename.lower().endswith(".pdf"):
            continue
        pdf_bytes = f.read()
        rec = process_pdf_bytes(pdf_bytes, f.filename)
        records.append(rec)

    return jsonify({"records": records, "fields": FIELDNAMES})


@app.route("/api/parse_one", methods=["POST"])
def api_parse_one():
    """逐文件解析接口（流式进度用）。"""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "未收到文件"}), 400
    pdf_bytes = f.read()
    rec = process_pdf_bytes(pdf_bytes, f.filename)
    return jsonify({"record": rec, "fields": FIELDNAMES})


@app.route("/api/download_csv", methods=["POST"])
def api_download_csv():
    """下载 CSV。"""
    data = request.get_json()
    records = data.get("records", [])

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=FIELDNAMES_EXT, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(records)

    csv_bytes = io.BytesIO(("\ufeff" + buf.getvalue()).encode("utf-8"))
    return send_file(
        csv_bytes,
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name="tickets.csv",
    )


@app.route("/api/download_excel", methods=["POST"])
def api_download_excel():
    """下载 Excel（.xlsx）- 多Sheet结构。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "服务端未安装 openpyxl，请联系管理员"}), 500

    data = request.get_json()
    records = data.get("records", [])

    wb = openpyxl.Workbook()
    
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14)
    thin = Side(style="thin", color="D9D9D9")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws1 = wb.active
    ws1.title = "行程明细"
    _write_detail_sheet(ws1, records, header_fill, header_font, cell_border)

    ws2 = wb.create_sheet("时间统计")
    _write_time_stats_sheet(ws2, records, header_fill, header_font, cell_border, title_font)

    ws3 = wb.create_sheet("路线分析")
    _write_route_stats_sheet(ws3, records, header_fill, header_font, cell_border, title_font)

    ws4 = wb.create_sheet("报销汇总")
    _write_reimb_stats_sheet(ws4, records, header_fill, header_font, cell_border, title_font)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="tickets.xlsx",
    )


def _write_detail_sheet(ws, records, header_fill, header_font, cell_border):
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    ws.append(FIELDNAMES_EXT)
    for col_idx, _ in enumerate(FIELDNAMES_EXT, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

    for row_idx, rec in enumerate(records, 2):
        for col_idx, field in enumerate(FIELDNAMES_EXT, 1):
            val = rec.get(field, "")
            if field == "票价(元)" and val:
                try:
                    val = float(val)
                except ValueError:
                    pass
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F5F8FF")

    for col_idx, field in enumerate(FIELDNAMES_EXT, 1):
        max_len = len(field)
        for row_idx in range(2, len(records) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def _write_time_stats_sheet(ws, records, header_fill, header_font, cell_border, title_font):
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    month_map = defaultdict(lambda: {"count": 0, "price": 0.0})
    quarter_map = defaultdict(lambda: {"count": 0, "price": 0.0})
    year_map = defaultdict(lambda: {"count": 0, "price": 0.0})

    for rec in records:
        date_str = rec.get("乘车日期", "")
        price = float(rec.get("票价(元)", 0) or 0)
        m = re.match(r"(\d{4})年(\d{1,2})月", date_str)
        if not m:
            continue
        year, month = int(m.group(1)), int(m.group(2))
        quarter = (month - 1) // 3 + 1
        
        month_key = f"{year}-{str(month).zfill(2)}"
        quarter_key = f"{year}Q{quarter}"
        year_key = str(year)
        
        month_map[month_key]["count"] += 1
        month_map[month_key]["price"] += price
        quarter_map[quarter_key]["count"] += 1
        quarter_map[quarter_key]["price"] += price
        year_map[year_key]["count"] += 1
        year_map[year_key]["price"] += price

    row = 1
    ws.cell(row=row, column=1, value="月度统计").font = title_font
    row += 1
    headers = ["月份", "出行次数", "费用合计(元)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border
    row += 1
    for k in sorted(month_map.keys()):
        ws.cell(row=row, column=1, value=k).border = cell_border
        ws.cell(row=row, column=2, value=month_map[k]["count"]).border = cell_border
        ws.cell(row=row, column=3, value=round(month_map[k]["price"], 2)).border = cell_border
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="季度统计").font = title_font
    row += 1
    for col, h in enumerate(["季度", "出行次数", "费用合计(元)"], 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border
    row += 1
    for k in sorted(quarter_map.keys()):
        ws.cell(row=row, column=1, value=k).border = cell_border
        ws.cell(row=row, column=2, value=quarter_map[k]["count"]).border = cell_border
        ws.cell(row=row, column=3, value=round(quarter_map[k]["price"], 2)).border = cell_border
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="年度统计").font = title_font
    row += 1
    for col, h in enumerate(["年度", "出行次数", "费用合计(元)"], 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border
    row += 1
    for k in sorted(year_map.keys()):
        ws.cell(row=row, column=1, value=k).border = cell_border
        ws.cell(row=row, column=2, value=year_map[k]["count"]).border = cell_border
        ws.cell(row=row, column=3, value=round(year_map[k]["price"], 2)).border = cell_border
        row += 1

    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _write_route_stats_sheet(ws, records, header_fill, header_font, cell_border, title_font):
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    route_map = defaultdict(lambda: {"count": 0, "price": 0.0})
    for rec in records:
        dep = rec.get("出发站", "")
        arr = rec.get("到达站", "")
        if not dep or not arr:
            continue
        key = f"{dep}→{arr}"
        route_map[key]["count"] += 1
        route_map[key]["price"] += float(rec.get("票价(元)", 0) or 0)

    sorted_routes = sorted(route_map.items(), key=lambda x: x[1]["price"], reverse=True)

    ws.cell(row=1, column=1, value="路线分析").font = title_font
    headers = ["路线", "出行次数", "费用合计(元)", "平均票价(元)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border

    for row_idx, (route, data) in enumerate(sorted_routes, 3):
        ws.cell(row=row_idx, column=1, value=route).border = cell_border
        ws.cell(row=row_idx, column=2, value=data["count"]).border = cell_border
        ws.cell(row=row_idx, column=3, value=round(data["price"], 2)).border = cell_border
        avg = round(data["price"] / data["count"], 2) if data["count"] > 0 else 0
        ws.cell(row=row_idx, column=4, value=avg).border = cell_border

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 20


def _write_reimb_stats_sheet(ws, records, header_fill, header_font, cell_border, title_font):
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    done_total = 0.0
    pending_total = 0.0
    done_count = 0
    pending_count = 0

    for rec in records:
        price = float(rec.get("票价(元)", 0) or 0)
        status = rec.get("报销状态", "未报销")
        if status == "已报销":
            done_total += price
            done_count += 1
        else:
            pending_total += price
            pending_count += 1

    ws.cell(row=1, column=1, value="报销汇总").font = title_font
    
    headers = ["状态", "票据数量", "费用合计(元)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border

    ws.cell(row=3, column=1, value="已报销").border = cell_border
    ws.cell(row=3, column=2, value=done_count).border = cell_border
    ws.cell(row=3, column=3, value=round(done_total, 2)).border = cell_border

    ws.cell(row=4, column=1, value="未报销").border = cell_border
    ws.cell(row=4, column=2, value=pending_count).border = cell_border
    ws.cell(row=4, column=3, value=round(pending_total, 2)).border = cell_border

    ws.cell(row=5, column=1, value="合计").border = cell_border
    ws.cell(row=5, column=1).font = Font(bold=True)
    ws.cell(row=5, column=2, value=done_count + pending_count).border = cell_border
    ws.cell(row=5, column=3, value=round(done_total + pending_total, 2)).border = cell_border

    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 16


# 兼容旧接口路径
@app.route("/api/download", methods=["POST"])
def api_download():
    return api_download_csv()


if __name__ == "__main__":
    app.run(debug=False, port=8088, host="127.0.0.1")