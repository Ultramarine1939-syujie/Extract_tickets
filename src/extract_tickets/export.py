"""CSV 与 Excel 导出。"""

from __future__ import annotations

import csv
import io
import re
from collections import defaultdict
from collections.abc import Iterable

from .constants import (
    DEFAULT_PENDING_REIMBURSEMENT,
    DONE_REIMBURSEMENT,
    FIELDNAMES_EXT,
    REIMBURSEMENT_FIELD,
)


def safe_float(value: object) -> float:
    """将票价字段转换为 float，无法转换时返回 0。"""

    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def build_csv_bytes(records: Iterable[dict]) -> io.BytesIO:
    """生成带 UTF-8 BOM 的 CSV 字节流。"""

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=FIELDNAMES_EXT, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(records)

    csv_bytes = io.BytesIO(("\ufeff" + buf.getvalue()).encode("utf-8"))
    csv_bytes.seek(0)
    return csv_bytes


def build_excel_bytes(records: list[dict]) -> io.BytesIO:
    """生成多 Sheet Excel 字节流。"""

    try:
        import openpyxl
        from openpyxl.styles import Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("服务端未安装 openpyxl，请联系管理员") from exc

    wb = openpyxl.Workbook()

    header_fill = PatternFill("solid", fgColor="1677FF")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14)
    thin = Side(style="thin", color="D9D9D9")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws1 = wb.active
    ws1.title = "行程明细"
    write_detail_sheet(ws1, records, header_fill, header_font, cell_border)

    ws2 = wb.create_sheet("时间统计")
    write_time_stats_sheet(ws2, records, header_fill, header_font, cell_border, title_font)

    ws3 = wb.create_sheet("路线分析")
    write_route_stats_sheet(ws3, records, header_fill, header_font, cell_border, title_font)

    ws4 = wb.create_sheet("报销汇总")
    write_reimb_stats_sheet(ws4, records, header_fill, header_font, cell_border, title_font)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def write_detail_sheet(ws, records, header_fill, header_font, cell_border) -> None:
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    ws.append(FIELDNAMES_EXT)
    for col_idx, _ in enumerate(FIELDNAMES_EXT, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

    for row_idx, rec in enumerate(records, 2):
        for col_idx, field in enumerate(FIELDNAMES_EXT, 1):
            value = rec.get(field, "")
            if field == "票价(元)" and value:
                value = safe_float(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F5F8FF")

    for col_idx, field in enumerate(FIELDNAMES_EXT, 1):
        max_len = len(field)
        for row_idx in range(2, len(records) + 2):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def write_time_stats_sheet(
    ws,
    records,
    header_fill,
    header_font,
    cell_border,
    title_font,
) -> None:
    from openpyxl.utils import get_column_letter

    month_map = defaultdict(lambda: {"count": 0, "price": 0.0})
    quarter_map = defaultdict(lambda: {"count": 0, "price": 0.0})
    year_map = defaultdict(lambda: {"count": 0, "price": 0.0})

    for rec in records:
        date_str = rec.get("乘车日期", "")
        price = safe_float(rec.get("票价(元)", 0))
        match = re.match(r"(\d{4})年(\d{1,2})月", date_str)
        if not match:
            continue
        year, month = int(match.group(1)), int(match.group(2))
        quarter = (month - 1) // 3 + 1

        month_key = f"{year}-{month:02d}"
        quarter_key = f"{year}Q{quarter}"
        year_key = str(year)

        month_map[month_key]["count"] += 1
        month_map[month_key]["price"] += price
        quarter_map[quarter_key]["count"] += 1
        quarter_map[quarter_key]["price"] += price
        year_map[year_key]["count"] += 1
        year_map[year_key]["price"] += price

    row = _write_stats_block(
        ws,
        1,
        "月度统计",
        ["月份", "出行次数", "费用合计(元)"],
        month_map,
        header_fill,
        header_font,
        cell_border,
        title_font,
    )
    row = _write_stats_block(
        ws,
        row + 2,
        "季度统计",
        ["季度", "出行次数", "费用合计(元)"],
        quarter_map,
        header_fill,
        header_font,
        cell_border,
        title_font,
    )
    _write_stats_block(
        ws,
        row + 2,
        "年度统计",
        ["年度", "出行次数", "费用合计(元)"],
        year_map,
        header_fill,
        header_font,
        cell_border,
        title_font,
    )

    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _write_stats_block(
    ws,
    start_row: int,
    title: str,
    headers: list[str],
    data: dict,
    header_fill,
    header_font,
    cell_border,
    title_font,
) -> int:
    row = start_row
    ws.cell(row=row, column=1, value=title).font = title_font
    row += 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border
    row += 1
    for key in sorted(data.keys()):
        ws.cell(row=row, column=1, value=key).border = cell_border
        ws.cell(row=row, column=2, value=data[key]["count"]).border = cell_border
        ws.cell(row=row, column=3, value=round(data[key]["price"], 2)).border = cell_border
        row += 1
    return row


def write_route_stats_sheet(ws, records, header_fill, header_font, cell_border, title_font) -> None:
    from openpyxl.utils import get_column_letter

    route_map = defaultdict(lambda: {"count": 0, "price": 0.0})
    for rec in records:
        dep = rec.get("出发站", "")
        arr = rec.get("到达站", "")
        if not dep or not arr:
            continue
        key = f"{dep}→{arr}"
        route_map[key]["count"] += 1
        route_map[key]["price"] += safe_float(rec.get("票价(元)", 0))

    sorted_routes = sorted(route_map.items(), key=lambda item: item[1]["price"], reverse=True)

    ws.cell(row=1, column=1, value="路线分析").font = title_font
    headers = ["路线", "出行次数", "费用合计(元)", "平均票价(元)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border

    for row_idx, (route, data) in enumerate(sorted_routes, 3):
        ws.cell(row=row_idx, column=1, value=route).border = cell_border
        ws.cell(row=row_idx, column=2, value=data["count"]).border = cell_border
        ws.cell(row=row_idx, column=3, value=round(data["price"], 2)).border = cell_border
        avg = round(data["price"] / data["count"], 2) if data["count"] else 0
        ws.cell(row=row_idx, column=4, value=avg).border = cell_border

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 20


def write_reimb_stats_sheet(ws, records, header_fill, header_font, cell_border, title_font) -> None:
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    done_total = pending_total = 0.0
    done_count = pending_count = 0

    for rec in records:
        price = safe_float(rec.get("票价(元)", 0))
        status = rec.get(REIMBURSEMENT_FIELD, DEFAULT_PENDING_REIMBURSEMENT)
        if status == DONE_REIMBURSEMENT:
            done_total += price
            done_count += 1
        else:
            pending_total += price
            pending_count += 1

    ws.cell(row=1, column=1, value="报销汇总").font = title_font

    headers = ["状态", "票据数量", "费用合计(元)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border

    rows = [
        (3, DONE_REIMBURSEMENT, done_count, round(done_total, 2)),
        (4, DEFAULT_PENDING_REIMBURSEMENT, pending_count, round(pending_total, 2)),
        (5, "合计", done_count + pending_count, round(done_total + pending_total, 2)),
    ]
    for row, label, count, total in rows:
        ws.cell(row=row, column=1, value=label).border = cell_border
        ws.cell(row=row, column=2, value=count).border = cell_border
        ws.cell(row=row, column=3, value=total).border = cell_border
    ws.cell(row=5, column=1).font = Font(bold=True)

    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 16
